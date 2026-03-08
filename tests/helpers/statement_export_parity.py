from __future__ import annotations

import csv
from dataclasses import dataclass
from decimal import Decimal, ROUND_HALF_UP
from io import BytesIO, StringIO
from typing import Any


REQUIRED_TOTAL_KEYS_BY_KIND: dict[str, tuple[str, ...]] = {
    "income": ("revenue", "expense", "net_income"),
    "balance": ("assets", "liabilities", "equity", "le_sum"),
    "cashflow": ("opening", "closing", "change", "operating", "investing", "financing", "net"),
}


@dataclass
class StatementExportParityResult:
    ok: bool
    selector_context: dict[str, Any]
    totals_mismatch_keys: list[str]
    metadata_mismatch_keys: list[str]
    status_mismatch_keys: list[str]
    message: str


def selector_context(selectors: dict[str, Any], kind: str) -> dict[str, Any]:
    return {
        "ym": selectors.get("ym"),
        "kind": kind,
        "ccy": selectors.get("ccy"),
        "cash_folders": selectors.get("cash_folders"),
        "source": selectors.get("source", "journal"),
    }


def quantize_cents(value: Any) -> Decimal:
    return Decimal(str(value)).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)


def _bool_text(value: Any) -> str:
    if isinstance(value, bool):
        return "true" if value else "false"
    if value is None:
        return ""
    return str(value).strip().lower()


def parse_csv_rows(raw_bytes: bytes) -> list[dict[str, str]]:
    text = raw_bytes.decode("utf-8")
    reader = csv.DictReader(StringIO(text))
    return [{str(k): ("" if v is None else str(v)) for k, v in row.items()} for row in reader]


def parse_xlsx_rows(raw_bytes: bytes) -> list[dict[str, str]]:
    try:
        from openpyxl import load_workbook
    except Exception as exc:
        raise RuntimeError("openpyxl is required for XLSX parity checks.") from exc

    workbook = load_workbook(filename=BytesIO(raw_bytes), read_only=True, data_only=True)
    sheet = workbook.active
    values = list(sheet.iter_rows(values_only=True))
    if not values:
        return []
    headers = ["" if value is None else str(value) for value in values[0]]
    rows: list[dict[str, str]] = []
    for row_values in values[1:]:
        row: dict[str, str] = {}
        for idx, header in enumerate(headers):
            value = row_values[idx] if idx < len(row_values) else None
            row[header] = "" if value is None else str(value)
        rows.append(row)
    return rows


def parse_export_rows(raw_bytes: bytes, fmt: str) -> list[dict[str, str]]:
    if fmt == "csv":
        return parse_csv_rows(raw_bytes)
    if fmt == "xlsx":
        return parse_xlsx_rows(raw_bytes)
    raise ValueError(f"Unsupported export format: {fmt}")


def _collect_export_totals(rows: list[dict[str, str]]) -> dict[str, Decimal]:
    out: dict[str, Decimal] = {}
    for row in rows:
        if (row.get("Section") or "") != "__TOTAL__":
            continue
        key = (row.get("Label") or "").strip()
        amount_raw = (row.get("amount_base") or "").strip()
        if not key or not amount_raw:
            continue
        out[key] = quantize_cents(amount_raw)
    return out


def _collect_export_meta_rows(rows: list[dict[str, str]]) -> dict[str, dict[str, str]]:
    out: dict[str, dict[str, str]] = {}
    for row in rows:
        if (row.get("Section") or "") != "__META__":
            continue
        key = (row.get("Label") or "").strip()
        if not key:
            continue
        out[key] = {
            "amount_base": (row.get("amount_base") or "").strip(),
            "meta_value": (row.get("meta_value") or "").strip(),
        }
    return out


def _build_message(
    selector_ctx: dict[str, Any],
    totals_mismatch_keys: set[str],
    metadata_mismatch_keys: set[str],
    status_mismatch_keys: set[str],
) -> str:
    return (
        "Statement export parity mismatch: "
        f"selector_context={selector_ctx} "
        f"totals_mismatch_keys={sorted(totals_mismatch_keys)} "
        f"metadata_mismatch_keys={sorted(metadata_mismatch_keys)} "
        f"status_mismatch_keys={sorted(status_mismatch_keys)}"
    )


def evaluate_statement_export_parity(
    *,
    selectors: dict[str, Any],
    kind: str,
    data_status_code: int,
    data_payload: dict[str, Any] | None,
    export_status_code: int,
    export_content_type: str | None = None,
    export_bytes: bytes,
    fmt: str,
    simulate_total_mismatch_key: str | None = None,
) -> StatementExportParityResult:
    selector_ctx = selector_context(selectors, kind)
    totals_mismatch_keys: set[str] = set()
    metadata_mismatch_keys: set[str] = set()
    status_mismatch_keys: set[str] = set()

    if data_status_code != export_status_code:
        status_mismatch_keys.add("status_code")

    if data_status_code != 200 or export_status_code != 200:
        message = _build_message(selector_ctx, totals_mismatch_keys, metadata_mismatch_keys, status_mismatch_keys)
        ok = not status_mismatch_keys
        return StatementExportParityResult(
            ok=ok,
            selector_context=selector_ctx,
            totals_mismatch_keys=sorted(totals_mismatch_keys),
            metadata_mismatch_keys=sorted(metadata_mismatch_keys),
            status_mismatch_keys=sorted(status_mismatch_keys),
            message="" if ok else message,
        )

    if not isinstance(data_payload, dict):
        status_mismatch_keys.add("data_payload_invalid")
        message = _build_message(selector_ctx, totals_mismatch_keys, metadata_mismatch_keys, status_mismatch_keys)
        return StatementExportParityResult(
            ok=False,
            selector_context=selector_ctx,
            totals_mismatch_keys=sorted(totals_mismatch_keys),
            metadata_mismatch_keys=sorted(metadata_mismatch_keys),
            status_mismatch_keys=sorted(status_mismatch_keys),
            message=message,
        )

    try:
        export_rows = parse_export_rows(export_bytes, fmt=fmt)
    except Exception as exc:
        status_mismatch_keys.add("export_parse_error")
        first8_hex = (export_bytes or b"")[:8].hex()
        first16_text = (export_bytes or b"")[:16].decode("utf-8", errors="replace")
        message = _build_message(selector_ctx, totals_mismatch_keys, metadata_mismatch_keys, status_mismatch_keys)
        message = (
            f"{message} "
            f"export_status_code={export_status_code} "
            f"content_type={export_content_type or ''} "
            f"first8_hex={first8_hex} "
            f"first16_text={first16_text!r} "
            f"parse_error={exc}"
        )
        return StatementExportParityResult(
            ok=False,
            selector_context=selector_ctx,
            totals_mismatch_keys=sorted(totals_mismatch_keys),
            metadata_mismatch_keys=sorted(metadata_mismatch_keys),
            status_mismatch_keys=sorted(status_mismatch_keys),
            message=message,
        )

    export_totals = _collect_export_totals(export_rows)
    export_meta = _collect_export_meta_rows(export_rows)

    required_total_keys = REQUIRED_TOTAL_KEYS_BY_KIND[kind]
    data_totals = (((data_payload.get("statements") or {}).get(kind) or {}).get("totals") or {})

    if simulate_total_mismatch_key:
        key = simulate_total_mismatch_key if simulate_total_mismatch_key in required_total_keys else required_total_keys[0]
        base = export_totals.get(key, Decimal("0.00"))
        export_totals[key] = quantize_cents(base + Decimal("0.01"))

    for key in required_total_keys:
        data_value = data_totals.get(key, None)
        if data_value is None:
            totals_mismatch_keys.add(key)
            continue
        if key not in export_totals:
            totals_mismatch_keys.add(key)
            continue
        if quantize_cents(data_value) != export_totals[key]:
            totals_mismatch_keys.add(key)

    source = data_payload.get("source") or {}
    expected_source_meta = {
        "source.mode": _bool_text(source.get("mode")),
        "source.mixed_mode_allowed": _bool_text(source.get("mixed_mode_allowed")),
        "source.legacy_rows_included_in_totals": _bool_text(source.get("legacy_rows_included_in_totals")),
    }
    for key, expected_value in expected_source_meta.items():
        actual = (export_meta.get(key) or {}).get("meta_value", "").strip().lower()
        if actual != expected_value:
            metadata_mismatch_keys.add(key)

    coverage = data_payload.get("coverage")
    if isinstance(coverage, dict):
        if (export_meta.get("coverage.present") or {}).get("meta_value", "").strip().lower() != "true":
            metadata_mismatch_keys.add("coverage.present")
        expected_numeric_coverage = {
            cov_key: cov_val
            for cov_key, cov_val in coverage.items()
            if isinstance(cov_val, (int, float, Decimal)) and not isinstance(cov_val, bool)
        }
        for cov_key, cov_val in expected_numeric_coverage.items():
            export_key = f"coverage.{cov_key}"
            if export_key not in export_meta:
                metadata_mismatch_keys.add(export_key)
                continue
            export_amount = (export_meta[export_key] or {}).get("amount_base", "")
            if not export_amount:
                metadata_mismatch_keys.add(export_key)
                continue
            if quantize_cents(cov_val) != quantize_cents(export_amount):
                metadata_mismatch_keys.add(export_key)
    else:
        if (export_meta.get("coverage.present") or {}).get("meta_value", "").strip().lower() != "false":
            metadata_mismatch_keys.add("coverage.present")
        for export_key in export_meta:
            if export_key.startswith("coverage.") and export_key != "coverage.present":
                metadata_mismatch_keys.add(export_key)

    ok = not totals_mismatch_keys and not metadata_mismatch_keys and not status_mismatch_keys
    message = ""
    if not ok:
        message = _build_message(selector_ctx, totals_mismatch_keys, metadata_mismatch_keys, status_mismatch_keys)

    return StatementExportParityResult(
        ok=ok,
        selector_context=selector_ctx,
        totals_mismatch_keys=sorted(totals_mismatch_keys),
        metadata_mismatch_keys=sorted(metadata_mismatch_keys),
        status_mismatch_keys=sorted(status_mismatch_keys),
        message=message,
    )
